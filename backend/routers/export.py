from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
import io
from database.db import get_db
from database import models
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, BarChart, Reference

router = APIRouter(prefix="/export", tags=["export"])

@router.get("/excel/{month_id}")
def export_excel(month_id: int, db: Session = Depends(get_db)):
    month = db.query(models.Month).filter(models.Month.id == month_id).first()
    if not month:
        raise HTTPException(status_code=404, detail="Month not found")

    expenses = db.query(models.Expense).filter(models.Expense.month_id == month_id).all()
    incomes = db.query(models.Income).filter(models.Income.month_id == month_id).all()
    budgets = db.query(models.Budget).filter(models.Budget.month_id == month_id).all()

    # Create DataFrames
    df_expenses = pd.DataFrame([{
        "Date": e.date,
        "Amount": e.amount,
        "Description": e.description,
        "Category": e.category,
        "Person": e.person
    } for e in expenses])

    df_incomes = pd.DataFrame([{
        "Date": i.date,
        "Amount": i.amount,
        "Description": i.description,
        "Category": i.category,
        "Person": i.person
    } for i in incomes])

    # Summary: Planned vs Actual
    summary_data = []

    # Expense summary
    expense_cats = {b.category: b.planned_amount for b in budgets if b.type == "expense"}
    actual_expenses = {}
    for e in expenses:
        actual_expenses[e.category] = actual_expenses.get(e.category, 0.0) + e.amount

    # Ensure all budget categories are included, plus any expense categories that have actuals but no budget
    all_expense_cats = set(expense_cats.keys()) | set(actual_expenses.keys())
    for cat in sorted(all_expense_cats):
        planned = expense_cats.get(cat, 0.0)
        actual = actual_expenses.get(cat, 0.0)
        summary_data.append({
            "Type": "Expense",
            "Category": cat,
            "Planned": planned,
            "Actual": actual,
            "Difference": planned - actual
        })

    # Income summary
    income_cats = {b.category: b.planned_amount for b in budgets if b.type == "income"}
    actual_incomes = {}
    for i in incomes:
        actual_incomes[i.category] = actual_incomes.get(i.category, 0.0) + i.amount

    all_income_cats = set(income_cats.keys()) | set(actual_incomes.keys())
    for cat in sorted(all_income_cats):
        planned = income_cats.get(cat, 0.0)
        actual = actual_incomes.get(cat, 0.0)
        summary_data.append({
            "Type": "Income",
            "Category": cat,
            "Planned": planned,
            "Actual": actual,
            "Difference": actual - planned
        })

    df_summary = pd.DataFrame(summary_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Summary Sheet
        df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=4)
        workbook = writer.book
        summary_sheet = writer.sheets['Summary']

        # Add Overall Header
        summary_sheet.cell(row=1, column=1, value=f"Budget Report: {month.month}/{month.year}")
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)

        total_income = sum(i.amount for i in incomes)
        total_expense = sum(e.amount for e in expenses)

        summary_sheet.cell(row=2, column=1, value="Total Income:")
        summary_sheet.cell(row=2, column=2, value=total_income)
        summary_sheet.cell(row=3, column=1, value="Total Expenses:")
        summary_sheet.cell(row=3, column=2, value=total_expense)
        summary_sheet.cell(row=3, column=3, value="Savings:")
        summary_sheet.cell(row=3, column=4, value=total_income - total_expense)

        # Style headers
        header_font = Font(bold=True)
        for col in range(1, df_summary.shape[1] + 1):
            cell = summary_sheet.cell(row=5, column=col)
            cell.font = header_font
            summary_sheet.column_dimensions[cell.column_letter].width = 20

        # Add Charts
        # 1. Pie Chart for Expenses
        if len(actual_expenses) > 0:
            pie = PieChart()
            labels = Reference(summary_sheet, min_col=2, min_row=6, max_row=5 + len(all_expense_cats))
            data = Reference(summary_sheet, min_col=4, min_row=6, max_row=5 + len(all_expense_cats))
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = "Expense Breakdown"
            summary_sheet.add_chart(pie, "G2")

        # 2. Bar Chart for Planned vs Actual
        bar = BarChart()
        data = Reference(summary_sheet, min_col=3, max_col=4, min_row=5, max_row=5 + len(summary_data))
        cats = Reference(summary_sheet, min_col=2, min_row=6, max_row=5 + len(summary_data))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.title = "Planned vs Actual"
        summary_sheet.add_chart(bar, "G18")

        # 2. Expenses Sheet
        df_expenses.to_excel(writer, sheet_name='Expenses', index=False)
        expenses_sheet = writer.sheets['Expenses']
        for col in range(1, df_expenses.shape[1] + 1):
            cell = expenses_sheet.cell(row=1, column=col)
            cell.font = header_font
            expenses_sheet.column_dimensions[cell.column_letter].width = 25

        # 3. Incomes Sheet
        df_incomes.to_excel(writer, sheet_name='Incomes', index=False)
        incomes_sheet = writer.sheets['Incomes']
        for col in range(1, df_incomes.shape[1] + 1):
            cell = incomes_sheet.cell(row=1, column=col)
            cell.font = header_font
            incomes_sheet.column_dimensions[cell.column_letter].width = 25

    output.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="budget_{month.year}_{month.month}.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.get("/csv/{month_id}")
def export_csv(month_id: int, db: Session = Depends(get_db)):
    # CSV will just export expenses for simplicity, or we could combine. 
    # Let's export expenses as it's the most common need for CSV.
    expenses = db.query(models.Expense).filter(models.Expense.month_id == month_id).all()
    df = pd.DataFrame([{
        "Date": e.date,
        "Amount": e.amount,
        "Description": e.description,
        "Category": e.category,
        "Person": e.person
    } for e in expenses])
    
    stream = io.StringIO()
    df.to_csv(stream, index=False)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=expenses_{month_id}.csv"
    return response
